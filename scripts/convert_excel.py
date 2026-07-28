from __future__ import annotations

import json
import re
import unicodedata
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "equipment.xlsx"
OUT = ROOT / "web-data" / "nems-data.json"


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def norm(value) -> str:
    text = clean(value).lower().replace("đ", "d")
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", text).strip()


def date_value(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return clean(value)


def row_dict(headers, row):
    return {clean(headers[i]): row[i] if i < len(row) else "" for i in range(len(headers)) if clean(headers[i])}


def val(row, aliases):
    keys = list(row.keys())
    nkeys = {k: norm(k) for k in keys}
    for alias in aliases:
        na = norm(alias)
        for k in keys:
            if nkeys[k] == na:
                return row[k]
    for alias in aliases:
        na = norm(alias)
        for k in keys:
            if na and na in nkeys[k]:
                return row[k]
    return ""


def read_rows(ws):
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []
    result = []
    for row in rows:
        item = row_dict(headers, row)
        item["__COL_K__"] = row[10] if len(row) > 10 else ""
        item["__HAS_DATA__"] = any(clean(v) for v in row)
        result.append(item)
    return result


def main():
    if not XLSX.exists():
        raise SystemExit("equipment.xlsx not found")

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    if "Equipment_Register" not in wb.sheetnames:
        raise SystemExit('Missing sheet "Equipment_Register"')
    if "Asset_Event_Log1" not in wb.sheetnames:
        raise SystemExit('Missing sheet "Asset_Event_Log1"')

    er = read_rows(wb["Equipment_Register"])
    lr = read_rows(wb["Asset_Event_Log1"])

    equipment = []
    for r in er:
        e = {
            "no": clean(val(r, ["No."])),
            "category": clean(val(r, ["Phân loại"])),
            "asset": clean(val(r, ["Mã tài sản"])),
            "assetName": clean(val(r, ["Tên Theo mã tài sản", "Tên theo mã tài sản"])),
            "name": clean(val(r, ["TÊN Name", "TÊN", "Name"])),
            "model": clean(val(r, ["MODEL", "Model"])),
            "serial": clean(val(r, ["Serial"])),
            "parameter": clean(val(r, ["THÔNG SỐ KỸ THUẬT Parameter", "THÔNG SỐ KỸ THUẬT"])),
            "qty": clean(val(r, ["SỐ LƯỢNG Quantity", "SỐ LƯỢNG"])),
            "manufacturer": clean(val(r, ["HÃNG SX Manufacturer", "HÃNG SX"])),
            "vendor": clean(val(r, ["NHÀ CUNG CẤP Vendor", "NHÀ CUNG CẤP"])),
            "year": date_value(val(r, ["Năm sản xuất"])),
            "position": clean(val(r, ["VỊ TRÍ LẮP ĐẶT Position", "VỊ TRÍ LẮP ĐẶT"])),
            "maintenanceCycle": clean(val(r, ["CHU KỲ BẢO DƯỠNG Maintenance cycle", "CHU KỲ BẢO DƯỠNG"])),
            "installationDate": date_value(val(r, ["NGÀY LẮP ĐẶT Installation date", "NGÀY LẮP ĐẶT"])),
            "accreditation": date_value(val(r, ["KIỂM ĐỊNH Accreditation", "KIỂM ĐỊNH"])),
            "calibration": date_value(val(r, ["Hiệu chuẩn Calibration", "Hiệu chuẩn"])),
            "note": clean(val(r, ["GHI CHÚ Note", "GHI CHÚ"])),
            "image": clean(val(r, ["Hình ảnh"])),
            "document": clean(val(r, ["Document"])),
        }
        if e["asset"] and e["name"]:
            equipment.append(e)

    events = []
    for r in lr:
        asset_auto = clean(val(r, ["Mã tài sản tự điền"]))
        name_auto = clean(val(r, ["Tên thiết bị đúng DS"]))
        model_auto = clean(val(r, ["Model tự điền"]))
        asset_name_auto = clean(val(r, ["Tên theo mã tài sản tự điền"]))
        serial_auto = clean(val(r, ["Serial tự điền"]))
        approval = clean(val(r, ["Trạng thái duyệt"]))
        approver = clean(val(r, ["Người duyệt"]))

        e = {
            "id": clean(val(r, ["ID"])),
            "date": date_value(val(r, ["Ngày nhận thông tin"])),
            "area": clean(val(r, ["Vị trí khu vực"])),
            "asset": asset_auto or clean(val(r, ["Mã tài sản"])),
            "assetName": asset_name_auto or clean(val(r, ["Tên theo mã tài sản"])),
            "name": name_auto or clean(val(r, ["Tên thiết bị"])),
            "model": model_auto or clean(val(r, ["Model"])),
            "serial": serial_auto or clean(val(r, ["Số Serial"])),
            "eventType": clean(val(r, ["Loại trạng thái/sự kiện"])),
            "description": clean(val(r, ["Nôi dung thông tin", "Nội dung thông tin"])),
            "technician": clean(val(r, ["Người phụ trách thực hiện"])),
            "result": clean(val(r, ["Kết quả"])),
            "columnK": clean(r.get("__COL_K__", "")),
            "downtime": clean(val(r, ["Downtime (Phút)", "Downtime"])),
            "cause": clean(val(r, ["Nguyên Nhân", "Nguyên nhân"])),
            "note": clean(val(r, ["Ghi chú"])),
            "approval": approval,
            "approver": approver if ("duyet" in norm(approval) or approver) else "",
        }
        # Giữ toàn bộ dòng lịch sử có dữ liệu, kể cả khi chưa có mã tài sản
        # hoặc tên thiết bị. Điều này giúp auditor xem đủ nhật ký theo ngày.
        if r.get("__HAS_DATA__") and any([e["date"], e["asset"], e["name"], e["eventType"], e["description"], e["columnK"], e["result"], e["note"]]):
            events.append(e)

    payload = {
        "generated": datetime.now(timezone(timedelta(hours=7))).strftime("%Y-%m-%d %H:%M:%S"),
        "equipment": equipment,
        "events": events,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print(f"Wrote {OUT}: {len(equipment)} equipment, {len(events)} events")


if __name__ == "__main__":
    main()
